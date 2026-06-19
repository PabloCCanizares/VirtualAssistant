"""Extraccion de texto desde bytes de documentos (PDF, CSV, DOCX, XLSX, TXT)."""

import csv
import logging
from io import BytesIO, StringIO

logger = logging.getLogger(__name__)

MAX_CSV_ROWS = 50
MAX_XLSX_ROWS = 50


def extract_text(raw_bytes: bytes, content_type: str, filename: str) -> str:
    """Devuelve texto legible extraido de los bytes del documento."""
    ext = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()

    try:
        if content_type == "application/pdf" or ext == "pdf":
            return _extract_pdf(raw_bytes)
        if ext == "csv" or content_type == "text/csv":
            return _extract_csv(raw_bytes)
        if ext == "docx" or "wordprocessingml" in (content_type or ""):
            return _extract_docx(raw_bytes)
        if ext == "xlsx" or "spreadsheetml" in (content_type or ""):
            return _extract_xlsx(raw_bytes)
        if ext in ("txt", "md", "json", "log") or (content_type or "").startswith("text/"):
            return raw_bytes.decode("utf-8", errors="replace")
        if (content_type or "").startswith("image/"):
            return "[Este documento es una imagen. No se puede extraer texto.]"
    except Exception:
        logger.exception("doc_parser: error extrayendo texto de %s", filename)
        return f"[Error al extraer texto del archivo '{filename}'.]"

    return f"[Formato no soportado para extraccion de texto: {ext or content_type}.]"


def _extract_pdf(raw_bytes: bytes) -> str:
    import pdfplumber

    parts = []
    with pdfplumber.open(BytesIO(raw_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                parts.append(page_text)
    return "\n\n".join(parts) or "[PDF sin texto extraible.]"


def _extract_csv(raw_bytes: bytes) -> str:
    content = raw_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(StringIO(content))
    rows = list(reader)
    if not rows:
        return "[CSV vacio.]"

    header = rows[0]
    lines = ["| " + " | ".join(header) + " |"]
    lines.append("| " + " | ".join(["---"] * len(header)) + " |")
    for row in rows[1 : MAX_CSV_ROWS + 1]:
        lines.append("| " + " | ".join(row) + " |")
    if len(rows) > MAX_CSV_ROWS + 1:
        lines.append(f"... ({len(rows) - MAX_CSV_ROWS - 1} filas adicionales omitidas)")
    return "\n".join(lines)


def _extract_docx(raw_bytes: bytes) -> str:
    from docx import Document

    doc = Document(BytesIO(raw_bytes))
    text = "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
    return text or "[Documento DOCX sin texto.]"


def _extract_xlsx(raw_bytes: bytes) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        parts.append(f"## Hoja: {sheet_name}")
        header = rows[0]
        parts.append("| " + " | ".join(str(c or "") for c in header) + " |")
        parts.append("| " + " | ".join(["---"] * len(header)) + " |")
        for row in rows[1 : MAX_XLSX_ROWS + 1]:
            parts.append("| " + " | ".join(str(c or "") for c in row) + " |")
        if len(rows) > MAX_XLSX_ROWS + 1:
            parts.append(f"... ({len(rows) - MAX_XLSX_ROWS - 1} filas adicionales omitidas)")
    wb.close()
    return "\n".join(parts) or "[Excel sin datos.]"
